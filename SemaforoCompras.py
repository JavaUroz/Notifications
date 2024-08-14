#-*- coding: utf-8 -*-
import pandas as pd
import os
from tokenize import Double
from xml.dom.minidom import TypeInfo
from xmlrpc.client import DateTime
import pyodbc
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import dotenv
from datetime import datetime
from datetime import timedelta
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill, Color, Fill


# Cargar librería para .env
dotenv.load_dotenv()

# Establecer el connection string
connection_string = os.environ['CONNECTION_STRING']

# Definir la consulta SQL
query = '''
WITH ImportesTot AS (
    SELECT
        SegDetC.sdcscc_ID,
        SegTotC.[stc_ImpNetoEmi] * (([stc_Tasa1] / 100) + 1) AS ImporteTot
    FROM SegDetC 
	INNER JOIN [SegTotC] ON [SegDetC].[sdcscc_ID] = [SegTotC].[stcscc_ID]
    GROUP BY SegDetC.sdcscc_ID, SegTotC.[stc_ImpNetoEmi] * (([stc_Tasa1] / 100) + 1)
)

SELECT --DISTINCT
	--[SegTiposC].[spcscc_ID],
    DATEDIFF(DAY, GETDATE(), [SegDetC].[sdc_FRecep]) AS [DIAS ENTREGA],
    CONVERT(varchar, [SegDetC].[sdc_FRecep], 103) AS [RECEPCION],
    ('(' + CONVERT(VARCHAR, ABS([SegCabC].[sccpro_Cod])) + ') ' + [SegCabC].[sccpro_RazSoc]) AS [PROVEEDOR],
    ([SegTiposC].[spctco_Cod] + ' ' + CONVERT(VARCHAR, ABS([SegTiposC].[spc_Nro]))) AS [COMPROBANTE],   
    CASE 
        WHEN [SegTiposC].[spctco_Cod] = 'OCR' THEN '_Reposición'
        ELSE
            CASE 
                WHEN ImportesTot.ImporteTot = 0 THEN '_No declarado'
                ELSE
                    CASE [SegCabC].[sccmon_codigo]
                        WHEN 1 THEN '$ '
                        WHEN 2 THEN 'U$S '
                    END + CONVERT(varchar(50), CAST(ImportesTot.ImporteTot AS decimal(38, 2)))
            END
    END AS [IMPORTE],

	CASE 
        WHEN [SegTiposC].[spctco_Cod] = 'OCR' THEN ''
        ELSE
            CASE 
                WHEN SUM((sdc_PrecioUn * sdc_CPendRtUM1) * (([stc_Tasa1] / 100) + 1)) = 0 THEN ''
                ELSE
                    CASE [SegCabC].[sccmon_codigo]
                        WHEN 1 THEN '$ '
                        WHEN 2 THEN 'U$S '
                    END + CONVERT(varchar(50), CAST(SUM((sdc_PrecioUn * sdc_CPendRtUM1) * (([stc_Tasa1] / 100) + 1)) AS decimal(38, 2)))
            END
    END AS [SALDO APROX],
    [SegCabC].[scc_Mens] AS [MENSAJE],
    [SegCabC].[scctxa_Texto] AS [OBSERVACIONES]
FROM [SBDACEST].[dbo].[SegTiposC]
INNER JOIN [SegDetC] ON [SegTiposC].[spcscc_ID] = [SegDetC].[sdcscc_ID]
INNER JOIN [SegCabC] ON [SegTiposC].[spcscc_ID] = [SegCabC].[scc_ID]
INNER JOIN [Proveed] ON [SegCabC].[sccpro_Cod] = [Proveed].[pro_Cod]
INNER JOIN [ImportesTot] ON [SegTiposC].[spcscc_ID] = [ImportesTot].[sdcscc_ID]
INNER JOIN [SegTotC] ON [SegTiposC].[spcscc_ID] = [SegTotC].[stcscc_ID]
WHERE 
    --sdcscc_ID = '170715' AND
    [sdc_TipoIt] != 'L' AND
    [spctco_Cod] != 'PC' AND
    ([sdc_CPendRtUM1] > 0 OR [sdc_CPendRtUM2] > 0) AND
    [spc_Nro] > 0 AND
    ([spctco_Cod] IN ('OC','OCP','OCR') OR
        ([spctco_Cod] = 'FC' AND [sccpro_Cod] IN ('008790', '010406'))
    ) AND 
    ([sdc_Desc] NOT IN ('Materiales para Fabricación',
                        'Materiales para Fabricación (IVA 10,5%)',
                        'Repuestos y Reparaciones (IVA 21%)',
                        'Ferias y Exposiciones (IVA 21%)',
                        'Materiales para la construcción',
                        'Ferretería - Artículos varios',
                        'Muebles y Utiles',
                        'Regalos Empresariales',
                        'Indumentaria',
                        'Reparaciones varias',
                        'Instalaciones (IVA 21%)',
                        'Gastos de Exposición',
                        'Mantenimiento Inmuebles (21%)',
                        'Fletes y Acarreos',
                        'Gastos Varios de Mantenimiento',
                        'Gastos de Seguridad e Higiene',
                        'Gastos de Fabricación',
                        'Publicidad (IVA 21%)') AND    
     ([sdc_Desc] NOT LIKE '%Materiales para Fabricación%' AND
      [sdc_Desc] NOT LIKE '%Repuestos y Reparaciones%' AND
      [sdc_Desc] NOT LIKE '%Ferias y Exposiciones%' AND
      [sdc_Desc] NOT LIKE '%Materiales para la construcción%' AND
      [sdc_Desc] NOT LIKE '%Ferretería - Artículos varios%' AND
      [sdc_Desc] NOT LIKE '%Muebles y Utiles%' AND
      [sdc_Desc] NOT LIKE '%mal facturada%')
    )
GROUP BY [SegTiposC].[spcscc_ID],
		 [SegDetC].[sdc_FRecep],
         [SegDetC].[sdc_FRecep],
         [SegCabC].[sccpro_Cod],
         [SegCabC].[sccpro_RazSoc],
         [SegTiposC].[spctco_Cod],
         [SegTiposC].[spc_Nro],
         [SegCabC].[scc_Mens],
         [SegCabC].[scctxa_Texto],
         [SegCabC].[sccmon_codigo],
         ImportesTot.ImporteTot
		 --((sdc_PrecioUn * sdc_CPendRtUM1) * (([stc_Tasa1] / 100) + 1))

ORDER BY [SegDetC].[sdc_FRecep]


'''

# Inicia la conexión
try:
    # Establecer la conexión con la base de datos
    conn = pyodbc.connect(connection_string)

    # Crear un cursor para ejecutar la consulta SQL
    cursor = conn.cursor()

    # Ejecutar la consulta SQL
    cursor.execute(query)

    # Obtener los resultados
    resultados = cursor.fetchall()
    
except pyodbc.Error as e:
    print('Ocurrio un error al conectar a la base de datos:', e)

# Ejecutar la consulta y cargar los resultados en un DataFrame de Pandas
df = pd.read_sql(query, conn)

# Cerrar el cursor y la conexión
cursor.close()
conn.close()


# # Mostrar los primeros registros del DataFrame para verificar los datos
# print(df.head())

# Ruta de Archivo Excel para exportar
ruta_archivo_excel = '//Mariana/compartidos/SEMAFORO COMPRAS.xlsx'
df.to_excel(ruta_archivo_excel, index=False)


# Cargar el archivo Excel con openpyxl
wb = load_workbook(ruta_archivo_excel)
sheet = wb.active

col_range = sheet.max_column

#Color rojo para cabecera
for col in range(1, col_range + 1):
        cell_header = sheet.cell(1, col)
        cell_header.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid") 


# ## ws is a openpypxl worksheet object
# _cell = sheet.cell('A1')

# # Font properties
# _cell.style.font.color.index = 'FFFFFFFF'
# _cell.style.font.name = 'Futura Lt BT'
# _cell.style.font.size = 11
# _cell.style.font.bold = True
# _cell.style.alignment.wrap_text = True



sheet.conditional_formatting.add('A2:A1000', ColorScaleRule(start_type='percentile', 
                                                       start_value=1, 
                                                       start_color='00FF0000', 
                                                       mid_type='percentile', 
                                                       mid_value=50, 
                                                       mid_color='00FFFF00', 
                                                       end_type='percentile', 
                                                       end_value=75, 
                                                       end_color='0000FF00'
                                                       )
                              )

# Ajustar el ancho de las columnas basado en el contenido de las celdas
for col in sheet.columns:
    max_length = 0
    column = col[0].column_letter  # Obtener la letra de la columna
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2  # Multiplicador para ajuste de ancho
    sheet.column_dimensions[column].width = adjusted_width

# Guardar los cambios en el archivo Excel
wb.save(ruta_archivo_excel)