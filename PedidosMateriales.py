# -*- coding: utf-8 -*-

import datetime
import os
import smtplib
import pyodbc
import pandas as pd
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.drawing.image import Image

# Cargar las variables de entorno desde el archivo .env
load_dotenv()

# Establecer el connection string desde las variables de entorno
connection_string = os.getenv('CONNECTION_STRING')

# Definir la consulta SQL
sql_query = '''
SELECT
    [SegTiposC].[spctco_Cod] AS COMP,
    ABS([SegTiposC].[spc_Nro]) AS NRO,
    CONVERT(varchar, [SegCabC].[scc_FEmision], 103) AS EMISION,
    CONVERT(varchar, MIN([SegDetC].[sdc_FRecep]),103) AS RECEPCION,
    SUBSTRING([SegDetC].[sdcart_CodGen], 1, 2) + '.' +
    SUBSTRING([SegDetC].[sdcart_CodGen], 3, 2) + '.' +
    SUBSTRING([SegDetC].[sdcart_CodGen], 5, 3) AS CODIGO,
    [SegDetC].[sdc_Desc] AS DESCRIPCION,
    [SegDetC].[sdcume_Desc1] AS UM1,    
    SUM([SegDetC].[sdc_CantUM1]) AS CANTIDAD,
    [SegDetC].[sdcume_Desc2] AS UM2,
    SUM([SegDetC].[sdc_CantUM2]) AS CANTIDAD
FROM [SBDACEST].[dbo].[SegTiposC]
INNER JOIN [SegDetC] ON [SegTiposC].[spcscc_ID] = [SegDetC].[sdcscc_ID]
FULL OUTER JOIN [SegCabC] ON [SegTiposC].[spcscc_ID] = [SegCabC].[scc_ID]
FULL OUTER JOIN [Proveed] ON [SegCabC].[sccpro_Cod] = [Proveed].[pro_Cod]
WHERE [SegTiposC].[spctco_Cod] = 'CPR' AND
    [SegCabC].[scc_FEmision] >= '2024-01-01 00:00:00.000' 
GROUP BY
    [SegTiposC].[spcscc_ID],
    [SegCabC].[sccpro_Cod],
    [SegCabC].[sccpro_RazSoc],
    [SegTiposC].[spctco_Cod],
    [SegTiposC].[spc_Nro],
    [SegCabC].[scc_FEmision],
    [SegDetC].[sdcart_CodGen],
    [SegDetC].[sdccon_Cod],
    [SegDetC].[sdc_Desc],
    [SegDetC].[sdcume_Desc1],
    [SegDetC].[sdcume_Desc2]
ORDER BY [SegTiposC].[spc_Nro], [SegCabC].[scc_FEmision], [SegDetC].[sdcart_CodGen];
'''

# Fecha y hora actuales
date_now = datetime.datetime.now().strftime('%d-%m-%Y_%H-%M-%S')

# Nombre de archivo Excel de salida principal
excel_file = 'Pedido_Materiales_CPR_{0}.xlsx'.format(date_now)
# excel_file = 'Pedido_Materiales_CPR_{0}.xlsx'.format(date_now.replace(':', '-'))

# Nombre de archivo Excel para los CPR exportados
exported_cprs_file = 'CPRExportadas.xlsx'

# Verificar y cargar los números de CPR ya exportados
existing_cprs = set()
if os.path.exists(exported_cprs_file):
    existing_df = pd.read_excel(exported_cprs_file)
    existing_cprs = set(existing_df['NRO'].astype(int))

# Establecer la conexión con la base de datos
conn = pyodbc.connect(connection_string)

# Ejecutar la consulta y obtener los datos en un DataFrame, filtrando los no exportados
df = pd.read_sql_query(sql_query, conn)
df = df[~df['NRO'].astype(int).isin(existing_cprs)]

# Cerrar la conexión
conn.close()

# Si no hay nuevos registros, salir del script
if df.empty:
    print('No hay nuevos registros para exportar.')
    exit()


df_rows = df.values.tolist()

# Crear un nuevo archivo Excel
wb = Workbook()
ws = wb.active

# Ruta de la imagen (suponiendo que está en la carpeta 'images' dentro del mismo directorio)
image_path = './images/logo-cestari-web.png'

# Crear un objeto Image de openpyxl
img = Image(image_path)

# Insertar la imagen en una celda específica (por ejemplo, A1)
ws.add_image(img, 'A1')

# Insertar 2 filas vacías al principio
ws.insert_rows(1, amount=2)

# Combinar celdas para el título PEDIDO DE MATERIALES CPR
title_cell = ws['F2']
title_cell.value = 'PEDIDO DE MATERIALES CPR'
title_cell.font = Font(name='Calibri', size=20, bold=True)
title_cell.alignment = Alignment(horizontal='center', vertical='center')

# Insertar 2 filas vacías después del título
ws.insert_rows(1, amount=1)

# Añadir el encabezado del DataFrame
header = list(df.columns)
ws.append(header)

# Aplicar formato de negrita al encabezado
for cell in ws[5]:
    cell.font = Font(size=12, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Añadir las filas del DataFrame
for row in df_rows:
    ws.append(row)

# Ajustar el ancho de las columnas basado en el contenido
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Obtener la letra de la columna
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2  # Ajuste para asegurar espacio adicional
    ws.column_dimensions[column].width = adjusted_width

# Definir bordes más gruesos para el encabezado
border_thick = Border(left=Side(style='medium'), 
                      right=Side(style='medium'), 
                      top=Side(style='medium'), 
                      bottom=Side(style='medium'))

# Definir bordes más finos para las celdas que no son del encabezado
border_thin = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# Aplicar bordes gruesos al rango de encabezados
for row in ws.iter_rows(min_row=5, max_row=5):
    for cell in row:
        cell.border = border_thick

# Aplicar bordes finos a las celdas de datos
for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
    for cell in row:
        cell.border = border_thin

# Aplicar color de relleno blanco a todas las celdas
for row in ws.iter_rows():
    for cell in row:
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# Guardar el archivo Excel
wb.save(excel_file)

# Actualizar el archivo de CPR exportados, evitando duplicados
exported_numbers = set(existing_df['NRO'].astype(int)) if 'NRO' in existing_df.columns else set()
exported_numbers.update(df['NRO'].astype(int))

# Guardar los números exportados en el archivo CPRExportadas.xlsx
exported_df = pd.DataFrame({'NRO': list(exported_numbers)})
exported_df.to_excel(exported_cprs_file, index=False, header=True)

# Configuracion del correo electronico
from_email = 'jvincent@imcestari.com'
to_email = 'mcelli@imcestari.com, javieruroz@imcestari.com'
subject = 'Pedido de Materiales por CPR'
body = 'Adjunto el archivo Excel con Pedido de Materiales.'

# Crear el mensaje
msg = MIMEMultipart()
msg['From'] = from_email
msg['To'] = to_email
msg['Subject'] = subject
msg.attach(MIMEText(body, 'plain'))

# Adjuntar el archivo Excel al mensaje
with open(excel_file, 'rb') as attachment:
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename= {excel_file}')
    msg.attach(part)

# Conectar al servidor SMTP y enviar el mensaje
username = os.environ['USUARIO']
password = os.environ['PASSWORD']

try:
    smtpObj = smtplib.SMTP_SSL('px000056.ferozo.com', 465)
    smtpObj.login(username, password)
    smtpObj.sendmail(from_email, to_email, msg.as_string())
    smtpObj.quit()
    print('Correo electronico enviado correctamente.')
except Exception as e:
    print(f'Error al enviar el correo electronico: {str(e)}')
